"""Export verified selection decisions as PDF and XLSX files."""

import re
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import (
    Flowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.tools.shopping_summary import SelectionReport, ShoppingSummaryOutput

_SAFE_BASENAME = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,127}$")
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_HEADERS = [
    "product_id",
    "title",
    "platform",
    "recommendation",
    "selling_price_cny",
    "landed_cost_cny",
    "total_cost_cny",
    "net_profit_cny",
    "profit_margin",
    "roi",
    "supplier_risk_level",
    "overall_score",
    "confidence",
    "reasons",
    "risks",
    "missing_data",
]


class ExportedReport(BaseModel):
    pdf_path: Path
    xlsx_path: Path


def _safe_cell(value: object) -> object:
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _row(item: SelectionReport) -> list[object]:
    return [
        _safe_cell(item.product_id),
        _safe_cell(item.title),
        item.platform,
        item.recommendation.value,
        item.selling_price_cny,
        item.landed_cost_cny,
        item.total_cost_cny,
        item.net_profit_cny,
        item.profit_margin,
        item.roi,
        item.supplier_risk_level.value if item.supplier_risk_level else None,
        item.overall_score,
        item.confidence,
        _safe_cell("; ".join(item.reasons)),
        _safe_cell("; ".join(item.risks)),
        _safe_cell("; ".join(item.missing_data)),
    ]


def _write_xlsx(summary: ShoppingSummaryOutput, target: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Selection Report"
    sheet.append(_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for report in summary.report:
        sheet.append(_row(report))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {"A": 18, "B": 38, "C": 14, "D": 16}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for column in range(5, len(_HEADERS) + 1):
        sheet.column_dimensions[chr(64 + column)].width = 18
    workbook.save(target)


def _write_pdf(summary: ShoppingSummaryOutput, target: Path) -> None:
    font_name = "STSong-Light"
    try:
        pdfmetrics.getFont(font_name)
    except KeyError:
        pdfmetrics.registerFont(UnicodeCIDFont(font_name))
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "LectorTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=18,
    )
    body_style = ParagraphStyle(
        "LectorBody",
        parent=styles["BodyText"],
        fontName=font_name,
        fontSize=9,
        leading=12,
        alignment=TA_LEFT,
    )
    document = SimpleDocTemplate(
        str(target), pagesize=landscape(A4), leftMargin=24, rightMargin=24
    )
    story: list[Flowable] = [
        Paragraph("Lector 选品报告", title_style),
        Spacer(1, 8),
        Paragraph(escape(summary.final_text), body_style),
        Spacer(1, 12),
    ]
    table_data: list[list[object]] = [
        [
            Paragraph("SKU", body_style),
            Paragraph("商品", body_style),
            Paragraph("结论", body_style),
            Paragraph("利润率", body_style),
            Paragraph("净利润 CNY", body_style),
            Paragraph("综合分", body_style),
            Paragraph("置信度", body_style),
            Paragraph("风险/缺失", body_style),
        ]
    ]
    for item in summary.report:
        risk_text = "; ".join([*item.risks, *item.missing_data]) or "无"
        table_data.append(
            [
                Paragraph(escape(item.product_id), body_style),
                Paragraph(escape(item.title), body_style),
                Paragraph(item.recommendation.value, body_style),
                item.profit_margin if item.profit_margin is not None else "-",
                item.net_profit_cny if item.net_profit_cny is not None else "-",
                item.overall_score,
                item.confidence,
                Paragraph(escape(risk_text), body_style),
            ]
        )
    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[60, 185, 70, 65, 75, 60, 60, 180],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(table)
    document.build(story)


def export_selection_report(
    summary: ShoppingSummaryOutput,
    output_dir: Path,
    basename: str = "selection-report",
) -> ExportedReport:
    """Write PDF and XLSX artifacts without recalculating business metrics."""
    if not _SAFE_BASENAME.fullmatch(basename):
        raise ValueError("basename must contain only safe filename characters")
    output_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = (output_dir / f"{basename}.pdf").resolve()
    xlsx_path = (output_dir / f"{basename}.xlsx").resolve()
    _write_pdf(summary, pdf_path)
    _write_xlsx(summary, xlsx_path)
    return ExportedReport(pdf_path=pdf_path, xlsx_path=xlsx_path)
